import streamlit as st
import pandas as pd
import openpyxl
import os
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import io
import tempfile
import numpy as np


# Set the page layout to wide
st.set_page_config(layout="wide")

# Title of the app
st.title(":bar_chart: Process Mapping & :hourglass: Cycle Time Simulation")

# Sidebar configuration
st.sidebar.header("Category")

# Create an expander for "Offers"
with st.sidebar.expander("Analysis"):
    # Create checkboxes for each offer
    new_analysis = st.checkbox("New")
    existing_analysis = st.checkbox("Existing")

# Display selected analysis
if new_analysis:
    st.subheader("New Analysis")

    # File uploader for the first Excel file (simulation_db.xlsx, sheet 'Process_CT')
    uploaded_file_simulation_db = st.file_uploader("Upload the simulation_db.xlsx file", type=["xlsx"])
    
    # Load data if the file is uploaded
    if uploaded_file_simulation_db:
        # Load the specific sheet from simulation_db.xlsx for 'Process_CT'
        df = pd.read_excel(uploaded_file_simulation_db, sheet_name='Process_CT')

        # Extract the required values
        shift_hr_day = df.at[0, 'Shift Hr/day']
        days_week = df.at[0, 'Days/Week']
        weeks_year = df.at[0, 'Weeks/Year']
        # hr_year_shift = df.at[0, 'Hr/Year (1 Shift)']
        hr_year_shift = shift_hr_day * days_week * weeks_year
        overall_labor_efficiency = df.at[0, 'Overall Labor Efficiency']
        # total_batch_setup_time = df.at[0, 'Total Batch Setup Time, sec']
        # total_cycle_time = df.at[0, 'Total Cycle Time, sec']
        
        # Hide the dataframe
        st.write("")

        # Create text inputs for each value
        col1, col2, col3 = st.columns(3)

        with col1:
            shift_hr_day_input = st.text_input('Shift Hr/day', value=shift_hr_day, disabled=True)
            # weeks_year_input = st.text_input('Weeks/Year', value=weeks_year, disabled=True)

        with col2:
            days_week_input = st.text_input('Days/Week', value=days_week, disabled=True)
            # hr_year_shift_input = st.text_input('Hr/Year (1 Shift)', value=hr_year_shift, disabled=True)

        with col3:
            overall_labor_efficiency_input = st.text_input('Overall Labor Efficiency', value=overall_labor_efficiency, disabled=True)

        # File uploader for the second Excel file (xydata.xlsx, sheet 'xydata_version')
        uploaded_file_xydata = st.file_uploader("Upload the xydata.xlsx file", type=["xlsx"])

        # File uploader for the third Excel file (simulation_db.xlsx, sheet 'SMD_Package_Feeder_Master')
        uploaded_file_feeder_master = st.file_uploader("Upload the feeder master Excel file (simulation_db.xlsx)", type=["xlsx"])

        # Load the data if both files are uploaded
        if uploaded_file_xydata and uploaded_file_feeder_master:
            # Load the specific sheet from xydata.xlsx
            df2 = pd.read_excel(uploaded_file_xydata, sheet_name='xydata_version')
            
            # Load the specific sheet from simulation_db.xlsx for feeder master
            feeder_master = pd.read_excel(uploaded_file_feeder_master, sheet_name='SMD_Package_Feeder_Master')
            
            # Perform VLOOKUP equivalent using merge
            df3 = df2.merge(feeder_master, left_on="Package", right_on="Package_Master", how="left")

            # Calculate values for the cycle time and other metrics
            component_count = df2['REFDES'].count()
            # total_cycle_time_calc = df3['Cycle Time_Master'].sum()
            bottom_pnp_cycle_time = df3[df3['Topbottom'] == 'NO']['Cycle Time_Master'].sum()
            top_pnp_cycle_time = df3[df3['Topbottom'] == 'YES']['Cycle Time_Master'].sum()

            xyoutput_cl1, xyoutput_cl2, xyoutput_cl3, xyoutput_cl4 = st.columns(4)              
            # Create text input boxes for the calculated values
            with xyoutput_cl1:
                solder_joints_input = st.text_input('Solder Joints')

            with xyoutput_cl2:
                component_count_input = st.text_input('Component Count', value=component_count, disabled=True)

            with xyoutput_cl3:
                bottom_pnp_cycle_time_input = st.text_input('Bottom Pick&Place Cycle Time', value=bottom_pnp_cycle_time, disabled=True)
                # total_batch_setup_time_input = st.text_input('Total Batch Setup Time, sec', value=total_batch_setup_time)

            with xyoutput_cl4:
                # total_cycle_time_input = st.text_input('Total Cycle Time, sec', value=total_cycle_time)
                top_pnp_cycle_time_input = st.text_input('Top Pick&Place Cycle Time', value=top_pnp_cycle_time, disabled=True)

        # Create an in-memory Excel file
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                # Write original 'xydata_version' sheet
                df2.to_excel(writer, sheet_name='xydata_version', index=False)

                # Create the 'Output' sheet explicitly
                writer.book.create_sheet(title="Output")
                output_sheet = writer.sheets["Output"]
                
                # Write merged 'Output' sheet
                output_sheet_name = 'Output'
                for r_idx, row in enumerate(dataframe_to_rows(df3, index=False, header=True), start=1):
                    for c_idx, value in enumerate(row, start=1):
                        writer.sheets[output_sheet_name].cell(row=r_idx, column=c_idx, value=value)

            # Save the in-memory file for download
            output.seek(0)

            # Generate the new filename based on the uploaded file name
            original_filename = uploaded_file_xydata.name
            file_root, file_ext = os.path.splitext(original_filename)
            edited_filename = f"{file_root}_edited_data{file_ext}"

            # Provide download button
            st.download_button(
                label="Download Updated xydata File",
                data=output,
                file_name=edited_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("Data has been successfully saved to the 'Output' sheet in the updated xydata file.")

            # Create an empty DataFrame with the defined columns
            initial_df = pd.DataFrame(columns=['Side', 'Stage', 'Batch Set up Time', 'Process Cycle Time'])

            # Initialize session state variables
            if 'data' not in st.session_state:
                st.session_state['data'] = initial_df

            if 'filtered_data' not in st.session_state:
                st.session_state['filtered_data'] = initial_df

            # Initialize dropdown values if not set
            if 'side' not in st.session_state:
                st.session_state['side'] = ''

            if 'stage' not in st.session_state:
                st.session_state['stage'] = ''

            if 'batch_setup_time' not in st.session_state:
                st.session_state['batch_setup_time'] = ''

            if 'process_cycle_time' not in st.session_state:
                st.session_state['process_cycle_time'] = ''

            if 'reset_selectbox' not in st.session_state:
                st.session_state['reset_selectbox'] = 0

            # Display the headings
            header_cols = st.columns(4)
            header_cols[0].markdown("<h6 style='text-align: center;'>Side</h6>", unsafe_allow_html=True)
            header_cols[1].markdown("<h6 style='text-align: center;'>Stage</h6>", unsafe_allow_html=True)
            header_cols[2].markdown("<h6 style='text-align: center;'>Batch Set up Time</h6>", unsafe_allow_html=True)
            header_cols[3].markdown("<h6 style='text-align: center;'>Process Cycle Time</h6>", unsafe_allow_html=True)

            # Function to display a row
            def display_row():
                row_cols = st.columns(4)
                
                # Select boxes to select the “Side” and “Stage”
                side = row_cols[0].selectbox('', [''] + list(df['Side'].unique()), key=f'side_{st.session_state.reset_selectbox}')
                stage = row_cols[1].selectbox('', [''] + list(df[df['Side'] == side]['Stage'].unique()) if side else [''], key=f'stage_{st.session_state.reset_selectbox}')
                
                # Display the values associated with “Stage” from keyvalue “Batch Set up Time” and “Process Cycle Time”
                batch_setup_time = df[(df['Side'] == side) & (df['Stage'] == stage)]['Batch Set up Time'].values[0] if side and stage else ''
                process_cycle_time = df[(df['Side'] == side) & (df['Stage'] == stage)]['Process Cycle Time'].values[0] if side and stage else ''

                with row_cols[2]:
                    batch_setup_time_input = st.text_input('', value=batch_setup_time, key=f'batch_setup_time_{st.session_state.reset_selectbox}')

                with row_cols[3]:
                    process_cycle_time_input = st.text_input('', value=process_cycle_time, key=f'process_cycle_time_{st.session_state.reset_selectbox}')

            # Display the row
            display_row()

            # Add Save, Clear, and Delete buttons
            save_col, clear_col, delete_col3, delete_col4 = st.columns(4)
            with save_col:
                if st.button('Save'):
                    # Save the current selection to session state data
                    side = st.session_state[f'side_{st.session_state.reset_selectbox}']
                    stage = st.session_state[f'stage_{st.session_state.reset_selectbox}']
                    batch_setup_time = st.session_state[f'batch_setup_time_{st.session_state.reset_selectbox}']
                    process_cycle_time = st.session_state[f'process_cycle_time_{st.session_state.reset_selectbox}']

                    if side and stage:
                        new_row = {
                            'Side': side,
                            'Stage': stage,
                            'Batch Set up Time': batch_setup_time,
                            'Process Cycle Time': process_cycle_time
                        }
                        if not ((st.session_state['filtered_data']['Side'] == side) & 
                                (st.session_state['filtered_data']['Stage'] == stage)).any():
                            st.session_state['filtered_data'] = pd.concat([st.session_state['filtered_data'], pd.DataFrame([new_row])], ignore_index=True)
                            st.success("Record added successfully. Select Your Next Side & Stage")
                        else:
                            st.warning("Record Already Exists in the Table")

            with clear_col:
                if st.button('Clear'):
                    # Increment the key to reset the select boxes
                    st.session_state['reset_selectbox'] += 1

            # Display the updated dataframe with a header
            st.markdown("## Process Mapping")
            st.dataframe(st.session_state['filtered_data'], use_container_width=True)
            edited_data=st.session_state['filtered_data']

            # Ensure the 'Process Cycle Time' & 'Batch Set up Time' column is numeric
            edited_data['Process Cycle Time'] = pd.to_numeric(edited_data['Process Cycle Time'], errors='coerce')
            edited_data['Batch Set up Time'] = pd.to_numeric(edited_data['Batch Set up Time'], errors='coerce')

            total_cycle_time_calc = edited_data['Process Cycle Time'].sum()
            total_batch_setup_time = edited_data['Batch Set up Time'].sum()


            # # Group by 'Side' and calculate the total sum of 'Process Cycle Time' for each side
            # side_cycle_time_sum = edited_data.groupby('Side')['Process Cycle Time'].sum()

            # Find the maximum overall Process Cycle Time
            Max_Overall_PCBA_CT_calc = edited_data['Process Cycle Time'].max()


            totalct_cl1, totalbst_cl2, xx_cl3, xx_cl4 = st.columns(4)              
            # Create text input boxes for the calculated values

            with totalct_cl1:
                total_cycle_time_display = st.text_input('Total Cycle Time', value=total_cycle_time_calc, disabled=True)

            with totalbst_cl2:
                total_batch_setup_time_display = st.text_input('Total Batch Set up Time', value=total_batch_setup_time, disabled=True)

            # Provide inputs for file name, sheet name, and path
            st.markdown("### Save Data to Excel")

            # Provide inputs for file name and sheet name only (without path)
            file_name = st.text_input("Enter the Excel file name (with .xlsx extension):")
            sheet_name = st.text_input("Enter the sheet name:")

            # Add a button to save the entire DataFrame
            if st.button("Save DataFrame to Excel"):
                with tempfile.TemporaryDirectory() as tmpdirname:
                    full_path = os.path.join(tmpdirname, file_name)

                    # Write data to Excel file
                    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                        # Prepare the DataFrame for saving
                        final_df = st.session_state['filtered_data'].copy()
                        
                        # Add the additional fields in the first row
                        for col, value in zip(
                            ['Max Overall PCBA CT', 'Shift Hr/day', 'Days/Week', 'Weeks/Year', 'Hr/Year (1 Shift)', 
                            'Overall Labor Efficiency', 'Total Batch Setup Time, sec', 'Total Cycle Time, sec', 
                            'Bottom P&P Cycle Time', 'Top P&P Cycle Time', 'Solder Joints', 'Component Count'],
                            [Max_Overall_PCBA_CT_calc, shift_hr_day, days_week, weeks_year, hr_year_shift, 
                            overall_labor_efficiency, total_batch_setup_time, total_cycle_time_calc, 
                            bottom_pnp_cycle_time, top_pnp_cycle_time, solder_joints_input, component_count]):
                            final_df.at[0, col] = value

                        final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Load the saved file and create a download link
                    with open(full_path, "rb") as f:
                        st.download_button(
                            label="Download Excel file",
                            data=f,
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
            # Add select box and delete button for deleting rows
            if not st.session_state['filtered_data'].empty:
                with delete_col3:
                    row_to_delete = st.selectbox('Select Row to Delete', st.session_state['filtered_data'].index + 1, key='row_to_delete')
                with delete_col4:
                    if st.button('Delete'):
                        # Delete the selected row
                        st.session_state['filtered_data'] = st.session_state['filtered_data'].drop(st.session_state['filtered_data'].index[row_to_delete - 1]).reset_index(drop=True)
                        st.rerun()
                        
                        # Save the updated DataFrame to the specified Excel sheet
                        full_path = os.path.join(save_path, file_name)
                        if os.path.exists(full_path):
                            with pd.ExcelWriter(full_path, engine='openpyxl', mode='a') as writer:
                                if sheet_name in writer.sheets:
                                    st.error(f"The sheet '{sheet_name}' already exists in the file '{file_name}'. Please choose a different sheet name.")
                                else:
                                    final_df = st.session_state['filtered_data'].copy()
                                    
                                    # Add the additional fields in the first row, without overwriting existing data
                                    for col, value in zip(
                                        ['Max Overall PCBA CT', 'Shift Hr/day', 'Days/Week', 'Weeks/Year', 'Hr/Year (1 Shift)', 
                                        'Overall Labor Efficiency', 'Total Batch Setup Time, sec', 'Total Cycle Time, sec', 
                                        'Bottom P&P Cycle Time', 'Top P&P Cycle Time', 'Solder Joints', 'Component Count'],
                                        [Max_Overall_PCBA_CT_calc, shift_hr_day, days_week, overall_labor_efficiency, 
                                        total_batch_setup_time, total_cycle_time_calc, bottom_pnp_cycle_time, top_pnp_cycle_time, solder_joints_input, 
                                        component_count]):
                                        final_df.at[0, col] = value

                                    final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                                    st.success(f"Updated DataFrame saved to {sheet_name} in {file_name}.")
                        else:
                            with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                                final_df = st.session_state['filtered_data'].copy()
                                
                                # Add the additional fields in the first row, without overwriting existing data
                                for col, value in zip(
                                    
                                    ['Max Overall PCBA CT', 'Shift Hr/day', 'Days/Week', 'Weeks/Year', 'Hr/Year (1 Shift)', 
                                    'Overall Labor Efficiency', 'Total Batch Setup Time, sec', 'Total Cycle Time, sec', 
                                    'Bottom P&P Cycle Time', 'Top P&P Cycle Time', 'Solder Joints', 'Component Count'],
                                    [Max_Overall_PCBA_CT_calc, shift_hr_day, days_week, weeks_year, hr_year_shift, 
                                    overall_labor_efficiency, total_batch_setup_time, total_cycle_time_calc, 
                                    bottom_pnp_cycle_time, top_pnp_cycle_time, solder_joints_input, component_count]):
                                    final_df.at[0, col] = value

                                final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                                st.success(f"Updated DataFrame saved to {sheet_name} in {file_name}.")

                        st.experimental_rerun()

##############################################################################################################################################
##############################################################################################################################################
##############################################################################################################################################
##############################################################################################################################################
##############################################################################################################################################
##############################################################################################################################################

if existing_analysis:
        st.subheader("Existing Analysis")

        # File uploader for Excel/CSV/XLSM files
        uploaded_file = st.file_uploader("Choose an Excel/CSV/XLSM file", type=["xlsx", "csv", "xlsm"])

        if uploaded_file:
            # Load data from the uploaded file
            @st.cache_data
            def load_data(file):
                if file.name.endswith('.csv'):
                    df = pd.read_csv(file)
                else:
                    df = pd.read_excel(file, sheet_name=None)
                return df

            data = load_data(uploaded_file)

            # Initialize session state to store edited data for each sheet
            if 'edited_sheets' not in st.session_state:
                st.session_state.edited_sheets = {}

            # Assuming 'data' is a dictionary of DataFrames loaded from an Excel file
            if isinstance(data, dict):
                sheet_name = st.selectbox("Select the sheet", data.keys())
                
                # Check if the sheet has been edited before; if so, load the edited version
                if sheet_name in st.session_state.edited_sheets:
                    st.session_state.df = st.session_state.edited_sheets[sheet_name]
                else:
                    selected_data = data[sheet_name]
                    st.session_state.df = pd.DataFrame(selected_data)  # Load original data from file
                
            # Extract the required values
            shift_hr_day = st.session_state.df.at[0, 'Shift Hr/day']
            days_week = st.session_state.df.at[0, 'Days/Week']
            weeks_year = st.session_state.df.at[0, 'Weeks/Year']
            hr_year_shift = st.session_state.df.at[0, 'Hr/Year (1 Shift)']
            overall_labor_efficiency = st.session_state.df.at[0, 'Overall Labor Efficiency']
            total_batch_setup_time = st.session_state.df.at[0, 'Total Batch Setup Time, sec']
            total_cycle_time = st.session_state.df.at[0, 'Total Cycle Time, sec']
            bottom_pnp_cycle_time = st.session_state.df.at[0, 'Bottom P&P Cycle Time']
            top_pnp_cycle_time = st.session_state.df.at[0, 'Top P&P Cycle Time']
            solder_joints = st.session_state.df.at[0, 'Solder Joints']
            component_count = st.session_state.df.at[0, 'Component Count']
            Max_Overall_PCBA_CT = st.session_state.df.at[0, 'Max Overall PCBA CT']
            
            # Hide the dataframe
            st.write("")

            # Create text inputs for each value
            col1, col2, col3 = st.columns(3)

            with col1:
                shift_hr_day_input = st.text_input('Shift Hr/day', value=shift_hr_day, disabled=True)
                weeks_year_input = st.text_input('Weeks/Year', value=weeks_year, disabled=True)
                total_cycle_time_input = st.text_input('Total Cycle Time, sec', value=total_cycle_time, disabled=True)
                Max_Overall_PCBA_CT_imput = st.text_input('Max Overall PCBA CT', value=Max_Overall_PCBA_CT, disabled=True)

            with col2:
                days_week_input = st.text_input('Days/Week', value=days_week, disabled=True)
                hr_year_shift_input = st.text_input('Hr/Year (1 Shift)', value=hr_year_shift, disabled=True)
                total_batch_setup_time_input = st.text_input('Total Batch Setup Time, sec', value=total_batch_setup_time, disabled=True)
                component_count_input = st.text_input('Component Count', value=component_count, disabled=True)

            with col3:
                overall_labor_efficiency_input = st.text_input('Overall Labor Efficiency', value=overall_labor_efficiency, disabled=True)
                solder_joints_input = st.text_input('Solder Joints', value=solder_joints, disabled=True)
                bottom_pnp_cycle_time_input = st.text_input('Bottom P&P Cycle Time', bottom_pnp_cycle_time, disabled=True)
                top_pnp_cycle_time_input = st.text_input('Top P&P Cycle Time', value=top_pnp_cycle_time, disabled=True)
            
            # Display data in a table
            st.subheader("Data Table")
            edited_data = st.data_editor(st.session_state.df)

            # Ensure the 'Process Cycle Time' & 'Batch Set up Time' column is numeric
            edited_data['Process Cycle Time'] = pd.to_numeric(edited_data['Process Cycle Time'], errors='coerce')
            edited_data['Batch Set up Time'] = pd.to_numeric(edited_data['Batch Set up Time'], errors='coerce')

            total_cycle_time_recalc = edited_data['Process Cycle Time'].sum()
            total_batch_setup_time_recalc = edited_data['Batch Set up Time'].sum()
            Max_Overall_PCBA_CT_recalc = edited_data['Process Cycle Time'].max()

            # Overwrite specific cells with the recalculated values
            edited_data.loc[0, 'Total Cycle Time, sec'] = total_cycle_time_recalc  # Example: overwrite row 0's 'Process Cycle Time'
            edited_data.loc[0, 'Total Batch Setup Time, sec'] = total_batch_setup_time_recalc  # Example: overwrite row 0's 'Batch Set up Time'
            edited_data.loc[0, 'Max Overall PCBA CT'] = Max_Overall_PCBA_CT_recalc  # Overwrite or add new column value

            # Save the updated DataFrame back to the session state
            st.session_state.edited_data = edited_data

            
            # Create buttons side by side for adding new rows and saving the table
            col1, col2 = st.columns([1, 1])

            # Create a separate save button for general edits in the table
            with col1:
                if st.button("Save Edited Table"):
                    # Save the edited data in the session state
                    st.session_state.edited_sheets[sheet_name] = edited_data

                    # Use BytesIO to create an in-memory buffer
                    output = io.BytesIO()
                    
                    # Write the edited data to this buffer
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        # Write each sheet in edited_sheets to the Excel writer
                        for sheet, df in st.session_state.edited_sheets.items():
                            df.to_excel(writer, sheet_name=sheet, index=False)
                    
                    # Move the pointer to the beginning of the BytesIO buffer
                    output.seek(0)

                    # Generate the new filename based on the uploaded file name
                    original_filename = uploaded_file.name
                    file_root, file_ext = os.path.splitext(original_filename)
                    edited_filename = f"{file_root}_edited_data{file_ext}"
                    
                    # Provide download button to download updated Excel file with the new file name
                    st.download_button(
                        label="Download Updated Excel File",
                        data=output,
                        file_name=edited_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.success("Table saved and download ready!")

            with col2:
                if st.button("Add New Row"):
                    # Create a new row with NaN values
                    new_row = pd.DataFrame({col: [np.nan] for col in edited_data.columns})
                    # Append the new row to the DataFrame
                    st.session_state.df = pd.concat([st.session_state.df, new_row], ignore_index=True)
                    st.session_state.edited_sheets[sheet_name] = st.session_state.df  # Update session state with new row
                    st.rerun()  # Rerun the app to update the data editor with the new row

            # Create buttons side by side for removing rows and saving removed rows
            col3, col4 = st.columns([1, 1])
            
            with col3:
                row_to_delete = st.selectbox("Select row to delete", st.session_state.df.index)
                if st.button("Remove Row"):
                    if 'removed_rows' not in st.session_state:
                        st.session_state.removed_rows = pd.DataFrame()

                    # Add the removed row to 'removed_rows' session state
                    st.session_state.removed_rows = pd.concat([st.session_state.removed_rows, st.session_state.df.loc[[row_to_delete]]])

                    # Drop the selected row and reset the index
                    st.session_state.df = st.session_state.df.drop(row_to_delete).reset_index(drop=True)
                    edited_data = st.session_state.df  # Ensure the edited data is synchronized
                    
                    # Recalculate values after the row is removed
                    total_cycle_time_recalc = edited_data['Process Cycle Time'].sum()
                    total_batch_setup_time_recalc = edited_data['Batch Set up Time'].sum()
                    Max_Overall_PCBA_CT_recalc = edited_data['Process Cycle Time'].max()
                    
                    # Overwrite specific cells with the recalculated values
                    edited_data.loc[0, 'Total Cycle Time, sec'] = total_cycle_time_recalc
                    edited_data.loc[0, 'Total Batch Setup Time, sec'] = total_batch_setup_time_recalc
                    edited_data.loc[0, 'Max Overall PCBA CT'] = Max_Overall_PCBA_CT_recalc

                    # Save the updated DataFrame back to the session state
                    st.session_state.edited_data = edited_data
                    st.session_state.edited_sheets[sheet_name] = edited_data  # Update session state with the new data
                    st.rerun()
            
            with col4:
                if st.button("Save Removed Rows"):
                    if 'removed_rows' in st.session_state and not st.session_state.removed_rows.empty:
                        # Use BytesIO to create an in-memory buffer
                        output = io.BytesIO()

                        # Create a dictionary of the Excel file with updated data in `st.session_state.df`
                        excel_file = pd.read_excel(uploaded_file, sheet_name=None)
                        excel_file[sheet_name] = st.session_state.df

                        # Write all sheets to this in-memory buffer
                        with pd.ExcelWriter(output, engine="openpyxl") as writer:
                            for sheet_name, sheet_data in excel_file.items():
                                sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)

                        # Move the pointer to the beginning of the BytesIO buffer
                        output.seek(0)

                        # Generate the new filename based on the uploaded file name
                        original_filename = uploaded_file.name
                        file_root, file_ext = os.path.splitext(original_filename)
                        removed_rows_filename = f"{file_root}_removed_rows{file_ext}"

                        # Provide download button to download updated Excel file with the new file name
                        st.download_button(
                            label="Download Excel File with Removed Rows",
                            data=output,
                            file_name=removed_rows_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.success("Removed rows saved and download ready!")

            st.subheader("Recalculated Values As Per Edited Data Table Values:")
            recalct_cl1, recalbst_cl2, recalmax_cl3 = st.columns(3)

            with recalct_cl1:
                st.text_input('Recalculated - Total Cycle Time, sec', value=total_cycle_time_recalc, disabled=True)
            with recalbst_cl2:
                st.text_input('Recalculated - Total Batch Setup Time, sec', value=total_batch_setup_time_recalc, disabled=True)
            with recalmax_cl3:
                st.text_input('Recalculated - Max Overall PCBA CT, sec', value=Max_Overall_PCBA_CT_recalc, disabled=True)
            
# revision history 16-Oct-24
# Added optoin to tempfile instead of os directory to handle cloud deployment in new_analysis
# existing_analysis is added along with new_analysis
# trial35_3.py - updated the layout and added dynamic calculation for Total CT, Max CT & Batch setup time.
