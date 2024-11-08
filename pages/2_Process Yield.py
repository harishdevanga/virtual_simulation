import streamlit as st
import pandas as pd
import openpyxl
import os
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Set the page layout to wide
st.set_page_config(layout="wide")

# Title of the app
st.title(":bar_chart: Process Mapping & :hourglass: Cycle Time Simulation")

# Sidebar configuration
st.sidebar.header("Category")

# Create an expander for "Offers"
with st.sidebar.expander("Analysis"):
    # Create checkboxes for each offer
    new_analysis = st.checkbox("New")
    existing_analysis = st.checkbox("Existing")

# Display selected analysis
if new_analysis:
    st.subheader("New Analysis")

    # File uploader for the first Excel file (simulation_db.xlsx, sheet 'Process_CT')
    uploaded_file_simulation_db = st.file_uploader("Upload the simulation_db.xlsx file", type=["xlsx"])
    
    # Load data if the file is uploaded
    if uploaded_file_simulation_db:
        # Load the specific sheet from simulation_db.xlsx for 'Process_CT'
        df = pd.read_excel(uploaded_file_simulation_db, sheet_name='Process_CT')

        # Extract the required values
        shift_hr_day = df.at[0, 'Shift Hr/day']
        days_week = df.at[0, 'Days/Week']
        weeks_year = df.at[0, 'Weeks/Year']
        hr_year_shift = df.at[0, 'Hr/Year (1 Shift)']
        overall_labor_efficiency = df.at[0, 'Overall Labor Efficiency']
        total_batch_setup_time = df.at[0, 'Total Batch Setup Time, sec']
        total_cycle_time = df.at[0, 'Total Cycle Time, sec']
        
        # Hide the dataframe
        st.write("")

        # Create text inputs for each value
        col1, col2, col3 = st.columns(3)

        with col1:
            shift_hr_day_input = st.text_input('Shift Hr/day', value=shift_hr_day, disabled=True)
            weeks_year_input = st.text_input('Weeks/Year', value=weeks_year, disabled=True)
            overall_labor_efficiency_input = st.text_input('Overall Labor Efficiency', value=overall_labor_efficiency, disabled=True)

        with col2:
            days_week_input = st.text_input('Days/Week', value=days_week, disabled=True)
            hr_year_shift_input = st.text_input('Hr/Year (1 Shift)', value=hr_year_shift, disabled=True)
            total_batch_setup_time_input = st.text_input('Total Batch Setup Time, sec', value=total_batch_setup_time)

        with col3:
            total_cycle_time_input = st.text_input('Total Cycle Time, sec', value=total_cycle_time)

        # File uploader for the second Excel file (xydata.xlsx, sheet 'xydata_version')
        uploaded_file_xydata = st.file_uploader("Upload the xydata.xlsx file", type=["xlsx"])

        # File uploader for the third Excel file (simulation_db.xlsx, sheet 'SMD_Package_Feeder_Master')
        uploaded_file_feeder_master = st.file_uploader("Upload the feeder master Excel file (simulation_db.xlsx)", type=["xlsx"])

        # Load the data if both files are uploaded
        if uploaded_file_xydata and uploaded_file_feeder_master:
            # Load the specific sheet from xydata.xlsx
            df2 = pd.read_excel(uploaded_file_xydata, sheet_name='xydata_version')
            
            # Load the specific sheet from simulation_db.xlsx for feeder master
            feeder_master = pd.read_excel(uploaded_file_feeder_master, sheet_name='SMD_Package_Feeder_Master')
            
            # Perform VLOOKUP equivalent using merge
            df3 = df2.merge(feeder_master, left_on="Package", right_on="Package_Master", how="left")

            # Calculate values for the cycle time and other metrics
            component_count = df2['REFDES'].count()
            total_cycle_time_calc = df3['Cycle Time_Master'].sum()
            bottom_cycle_time = df3[df3['Topbottom'] == 'NO']['Cycle Time_Master'].sum()
            top_cycle_time = df3[df3['Topbottom'] == 'YES']['Cycle Time_Master'].sum()

            # Create text input boxes for the calculated values
            with col1:
                solder_joints_input = st.text_input('Solder Joints')

            with col2:
                component_count_input = st.text_input('Component Count', value=component_count, disabled=True)

            with col3:
                bottom_cycle_time_input = st.text_input('Bottom Cycle Time', value=bottom_cycle_time, disabled=True)
                top_cycle_time_input = st.text_input('Top Cycle Time', value=top_cycle_time, disabled=True)

            # Save merged data to 'Output' sheet in the same workbook
            temp_path = 'D:/python/working_folder/ProcessMap&CTSimulation/xydata.xlsx'
            with open(temp_path, "wb") as f:
                f.write(uploaded_file_xydata.getbuffer())

            # Open the workbook and work with the 'Output' sheet
            wb = openpyxl.load_workbook(temp_path)

            # Check if the 'Output' sheet already exists
            if 'Output' in wb.sheetnames:
                output_sheet = wb['Output']
                # Clear the contents of the existing sheet by specifying a valid range
                max_col_letter = get_column_letter(output_sheet.max_column)  # Convert column index to letter
                max_row = output_sheet.max_row
                for row in output_sheet[f"A1:{max_col_letter}{max_row}"]:
                    for cell in row:
                        cell.value = None
            else:
                # Create a new sheet named 'Output'
                output_sheet = wb.create_sheet('Output')

            # Write df3 to 'Output' starting from cell A1
            for r_idx, row in enumerate(dataframe_to_rows(df3, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    output_sheet.cell(row=r_idx, column=c_idx, value=value)

            # Save the updated workbook
            wb.save(temp_path)
            st.success("Data has been successfully saved to the 'Output' sheet in xydata.xlsx.")

            # # Write the merged DataFrame to the sheet
            # new_sht.range("A1").options(index=False).value = df3
            # new_sht.range("J1:L1").color = (0, 204, 0)

            # Create an empty DataFrame with the defined columns
            initial_df = pd.DataFrame(columns=['Side', 'Stage', 'Batch Set up Time', 'Process Cycle Time'])

            # Initialize session state variables
            if 'data' not in st.session_state:
                st.session_state['data'] = initial_df

            if 'filtered_data' not in st.session_state:
                st.session_state['filtered_data'] = initial_df

            # Initialize dropdown values if not set
            if 'side' not in st.session_state:
                st.session_state['side'] = ''

            if 'stage' not in st.session_state:
                st.session_state['stage'] = ''

            if 'batch_setup_time' not in st.session_state:
                st.session_state['batch_setup_time'] = ''

            if 'process_cycle_time' not in st.session_state:
                st.session_state['process_cycle_time'] = ''

            if 'reset_selectbox' not in st.session_state:
                st.session_state['reset_selectbox'] = 0

            # Display the headings
            header_cols = st.columns(4)
            header_cols[0].markdown("<h6 style='text-align: center;'>Side</h6>", unsafe_allow_html=True)
            header_cols[1].markdown("<h6 style='text-align: center;'>Stage</h6>", unsafe_allow_html=True)
            header_cols[2].markdown("<h6 style='text-align: center;'>Batch Set up Time</h6>", unsafe_allow_html=True)
            header_cols[3].markdown("<h6 style='text-align: center;'>Process Cycle Time</h6>", unsafe_allow_html=True)

            # Function to display a row
            def display_row():
                row_cols = st.columns(4)
                
                # Select boxes to select the “Side” and “Stage”
                side = row_cols[0].selectbox('', [''] + list(df['Side'].unique()), key=f'side_{st.session_state.reset_selectbox}')
                stage = row_cols[1].selectbox('', [''] + list(df[df['Side'] == side]['Stage'].unique()) if side else [''], key=f'stage_{st.session_state.reset_selectbox}')
                
                # Display the values associated with “Stage” from keyvalue “Batch Set up Time” and “Process Cycle Time”
                batch_setup_time = df[(df['Side'] == side) & (df['Stage'] == stage)]['Batch Set up Time'].values[0] if side and stage else ''
                process_cycle_time = df[(df['Side'] == side) & (df['Stage'] == stage)]['Process Cycle Time'].values[0] if side and stage else ''

                with row_cols[2]:
                    batch_setup_time_input = st.text_input('', value=batch_setup_time, key=f'batch_setup_time_{st.session_state.reset_selectbox}')

                with row_cols[3]:
                    process_cycle_time_input = st.text_input('', value=process_cycle_time, key=f'process_cycle_time_{st.session_state.reset_selectbox}')

            # Display the row
            display_row()

            # Add Save, Clear, and Delete buttons
            save_col, clear_col, delete_col3, delete_col4 = st.columns(4)
            with save_col:
                if st.button('Save'):
                    # Save the current selection to session state data
                    side = st.session_state[f'side_{st.session_state.reset_selectbox}']
                    stage = st.session_state[f'stage_{st.session_state.reset_selectbox}']
                    batch_setup_time = st.session_state[f'batch_setup_time_{st.session_state.reset_selectbox}']
                    process_cycle_time = st.session_state[f'process_cycle_time_{st.session_state.reset_selectbox}']

                    if side and stage:
                        new_row = {
                            'Side': side,
                            'Stage': stage,
                            'Batch Set up Time': batch_setup_time,
                            'Process Cycle Time': process_cycle_time
                        }
                        if not ((st.session_state['filtered_data']['Side'] == side) & 
                                (st.session_state['filtered_data']['Stage'] == stage)).any():
                            st.session_state['filtered_data'] = pd.concat([st.session_state['filtered_data'], pd.DataFrame([new_row])], ignore_index=True)
                            st.success("Record added successfully. Select Your Next Side & Stage")
                        else:
                            st.warning("Record Already Exists in the Table")

            with clear_col:
                if st.button('Clear'):
                    # Increment the key to reset the select boxes
                    st.session_state['reset_selectbox'] += 1

            # Display the updated dataframe with a header
            st.markdown("## Process Mapping")
            st.dataframe(st.session_state['filtered_data'])

            # Provide inputs for file name, sheet name, and path
            st.markdown("### Save Data to Excel")

            file_name = st.text_input("Enter the Excel file name (with .xlsx extension):")
            sheet_name = st.text_input("Enter the sheet name:")
            save_path = st.text_input("Enter the path to save the Excel file:")

            # Add a button to save the entire DataFrame
            if st.button("Save DataFrame to Excel"):
                full_path = os.path.join(save_path, file_name)

                # Check if the file and sheet already exist
                if os.path.exists(full_path):
                    with pd.ExcelWriter(full_path, engine='openpyxl', mode='a') as writer:
                        if sheet_name in writer.sheets:
                            st.error(f"The sheet '{sheet_name}' already exists in the file '{file_name}'. Please choose a different sheet name.")
                        else:
                            # Prepare the DataFrame for saving
                            final_df = st.session_state['filtered_data'].copy()
                            
                            # Add the additional fields in the first row, without overwriting existing data
                            for col, value in zip(
                                ['Max Overall PCBA CT', 'Shift Hr/day', 'Days/Week', 'Weeks/Year', 'Hr/Year (1 Shift)', 
                                'Overall Labor Efficiency', 'Total Batch Setup Time, sec', 'Total Cycle Time, sec', 
                                'Bottom Cycle Time', 'Top Cycle Time', 'Solder Joints', 'Component Count'],
                                [total_cycle_time_calc, shift_hr_day, days_week, weeks_year, hr_year_shift, overall_labor_efficiency, 
                                total_batch_setup_time, total_cycle_time_calc, bottom_cycle_time, top_cycle_time, solder_joints_input, 
                                component_count]):
                                final_df.at[0, col] = value

                            final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            st.success(f"DataFrame saved successfully to {sheet_name} in {file_name}.")
                else:
                    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                        # Prepare the DataFrame for saving
                        final_df = st.session_state['filtered_data'].copy()
                        
                        # Add the additional fields in the first row, without overwriting existing data
                        for col, value in zip(
                            ['Max Overall PCBA CT', 'Shift Hr/day', 'Days/Week', 'Weeks/Year', 'Hr/Year (1 Shift)', 
                            'Overall Labor Efficiency', 'Total Batch Setup Time, sec', 'Total Cycle Time, sec', 
                            'Bottom Cycle Time', 'Top Cycle Time', 'Solder Joints', 'Component Count'],
                            [total_cycle_time_calc, shift_hr_day, days_week, weeks_year, hr_year_shift, overall_labor_efficiency, 
                            total_batch_setup_time, total_cycle_time_calc, bottom_cycle_time, top_cycle_time, solder_joints_input, 
                            component_count]):
                            final_df.at[0, col] = value

                        final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        st.success(f"DataFrame saved successfully to {sheet_name} in {file_name}.")

            # Add select box and delete button for deleting rows
            if not st.session_state['filtered_data'].empty:
                with delete_col3:
                    row_to_delete = st.selectbox('Select Row to Delete', st.session_state['filtered_data'].index + 1, key='row_to_delete')
                with delete_col4:
                    if st.button('Delete'):
                        # Delete the selected row
                        st.session_state['filtered_data'] = st.session_state['filtered_data'].drop(st.session_state['filtered_data'].index[row_to_delete - 1]).reset_index(drop=True)
                        st.rerun()
                        
                        # Save the updated DataFrame to the specified Excel sheet
                        full_path = os.path.join(save_path, file_name)
                        if os.path.exists(full_path):
                            with pd.ExcelWriter(full_path, engine='openpyxl', mode='a') as writer:
                                if sheet_name in writer.sheets:
                                    st.error(f"The sheet '{sheet_name}' already exists in the file '{file_name}'. Please choose a different sheet name.")
                                else:
                                    final_df = st.session_state['filtered_data'].copy()
                                    
                                    # Add the additional fields in the first row, without overwriting existing data
                                    for col, value in zip(
                                        ['Max Overall PCBA CT', 'Shift Hr/day', 'Days/Week', 'Weeks/Year', 'Hr/Year (1 Shift)', 
                                        'Overall Labor Efficiency', 'Total Batch Setup Time, sec', 'Total Cycle Time, sec', 
                                        'Bottom Cycle Time', 'Top Cycle Time', 'Solder Joints', 'Component Count'],
                                        [total_cycle_time_calc, shift_hr_day, days_week, weeks_year, hr_year_shift, overall_labor_efficiency, 
                                        total_batch_setup_time, total_cycle_time_calc, bottom_cycle_time, top_cycle_time, solder_joints_input, 
                                        component_count]):
                                        final_df.at[0, col] = value

                                    final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                                    st.success(f"Updated DataFrame saved to {sheet_name} in {file_name}.")
                        else:
                            with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                                final_df = st.session_state['filtered_data'].copy()
                                
                                # Add the additional fields in the first row, without overwriting existing data
                                for col, value in zip(
                                    ['Max Overall PCBA CT', 'Shift Hr/day', 'Days/Week', 'Weeks/Year', 'Hr/Year (1 Shift)', 
                                    'Overall Labor Efficiency', 'Total Batch Setup Time, sec', 'Total Cycle Time, sec', 
                                    'Bottom Cycle Time', 'Top Cycle Time', 'Solder Joints', 'Component Count'],
                                    [total_cycle_time_calc, shift_hr_day, days_week, weeks_year, hr_year_shift, overall_labor_efficiency, 
                                    total_batch_setup_time, total_cycle_time_calc, bottom_cycle_time, top_cycle_time, solder_joints_input, 
                                    component_count]):
                                    final_df.at[0, col] = value

                                final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                                st.success(f"Updated DataFrame saved to {sheet_name} in {file_name}.")

                        st.experimental_rerun()
